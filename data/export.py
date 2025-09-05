from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file

import io
from data.formatacao import formato_brasileiro


def exportar_excel(resultados, nf_cache):
    if not resultados:
        return "Nenhum documento para exportar", 400

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão

    for nome, conteudo in resultados.items():
        dados = conteudo["dados"]
        despesas = conteudo["despesas"]

        ws = wb.create_sheet(title=nome[:31])

        # Cores e estilos
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")
        success_fill = PatternFill(
            start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        danger_fill = PatternFill(
            start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        warning_fill = PatternFill(
            start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # -----------------------------
        # 1) Cabeçalho com os "cards"
        # -----------------------------
        ws.append(["Documento:", nome])
        ws.append(["Importador/Exportador:",
                  dados.get("Importador/Exportador", "")])
        ws.append(["Referência:", dados.get("Referência", "")])
        ws.append(["Total não Trib:", dados.get("Total não Trib.", "0")])

        # Formatação do cabeçalho
        for row in range(1, 5):
            ws[f'A{row}'].font = bold_font
            ws[f'A{row}'].fill = warning_fill
            ws[f'B{row}'].border = thin_border
            ws[f'A{row}'].border = thin_border

        # -----------------------------
        # 2) Totais (aplicando as mesmas lógicas do HTML)
        # -----------------------------
        total_sem_frete = 0
        total_liquido = 0

        for idx, d in enumerate(despesas):
            valor_int = d.get("valorInt", 0)
            descricao = d.get("Descrição", "")

            # Lógica do valor líquido igual ao HTML
            if descricao == "FRETE INTERNACIONAL":
                valor_liquido_item = 0  # Valor líquido zero para frete internacional
            else:
                valor_liquido_item = valor_int
                total_sem_frete += valor_int

            total_liquido += valor_liquido_item

        formula_check = total_liquido - total_sem_frete

        ws.append([])  # linha vazia
        ws.append(["Total sem Frete", formato_brasileiro(total_sem_frete)])
        ws.append(["Total Líquido", formato_brasileiro(total_liquido)])
        ws.append(["Check (Líquido - Sem Frete)",
                  formato_brasileiro(formula_check)])
        ws.append([])  # linha vazia

        # Formatação dos totais
        for row in range(6, 9):
            ws[f'A{row}'].font = bold_font
            ws[f'A{row}'].fill = success_fill if row == 8 else warning_fill
            ws[f'B{row}'].border = thin_border
            ws[f'A{row}'].border = thin_border

        # -----------------------------
        # 3) Cabeçalho da Tabela de Despesas
        # -----------------------------
        header_row = 10
        headers = [
            "Descrição", "Valor", "NF Filha", "NF Mãe", "Nº DOC",
            "CFOP", "Parceiro", "Processo", "Conta Contábil",
            "Valor Líquido", "ICMS %", "Valor Bruto",
            "Classificação", "PIS", "COFINS", "ICMS", "Emissão de NFe"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # -----------------------------
        # 4) Dados da Tabela (aplicando todas as lógicas do HTML)
        # -----------------------------
        for idx, d in enumerate(despesas):
            row = header_row + idx + 1
            nf_filha_valor = nf_cache.get(nome, {}).get(str(idx), "")

            valor_int = d.get("valorInt", 0)
            descricao = d.get("Descrição", "")
            classificacao = d.get("Classificação", "")
            icms_percent = dados.get("ICMS%", "")

            # Aplicando as mesmas lógicas do HTML

            # Valor Líquido (coluna J)
            if descricao == "FRETE INTERNACIONAL":
                valor_liquido = "0,00"
                liquido_cell_fill = danger_fill
            else:
                valor_liquido = formato_brasileiro(d.get("Valor", ""))
                liquido_cell_fill = None

            # Valor Bruto/ICMS (coluna L e P)
            if descricao == "FRETE INTERNACIONAL":
                valor_bruto = "0,00"
                icms_valor = "0,00"
            elif icms_percent == "18.00%":
                icms_calculado = valor_int * 0.18
                valor_bruto = formato_brasileiro(icms_calculado)
                icms_valor = formato_brasileiro(icms_calculado)
            else:
                valor_bruto = "0,00"
                icms_valor = "0,00"

            # PIS e COFINS (colunas N e O)
            if classificacao == "COM CRÉDITO":
                pis = formato_brasileiro(valor_int * 0.0175)
                cofins = formato_brasileiro(valor_int * 0.076)
                pis_fill = success_fill
                cofins_fill = success_fill
            else:
                pis = "0,00"
                cofins = "0,00"
                pis_fill = None
                cofins_fill = None

            # Emissão de NFe
            emissao_nfe = "COM EMISSÃO" if nf_filha_valor else "SEM EMISSÃO"
            emissao_fill = success_fill if nf_filha_valor else danger_fill

            # Preenchendo a linha
            row_data = [
                descricao,                              # A - Descrição
                valor_int,                              # B - Valor (numérico)
                nf_filha_valor,                        # C - NF Filha
                dados.get("NFMAE", ""),               # D - NF Mãe
                dados.get("Nº DOC", ""),              # E - Nº DOC
                dados.get("CFOP", ""),                # F - CFOP
                dados.get("PARCEIRO", ""),            # G - Parceiro
                dados.get("Referência", ""),          # H - Processo
                dados.get("CONTABIL", ""),            # I - Conta Contábil
                # J - Valor Líquido (numérico)
                valor_liquido,
                icms_percent,                          # K - ICMS %
                # L - Valor Bruto (numérico)
                valor_bruto,
                classificacao,                         # M - Classificação
                pis,                                  # N - PIS (numérico)
                cofins,                               # O - COFINS (numérico)
                icms_valor,                           # P - ICMS (numérico)
                emissao_nfe                           # Q - Emissão de NFe
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.border = thin_border

                # Formatação numérica para colunas de valores
                # Valor, Valor Líquido, Valor Bruto, PIS, COFINS, ICMS
                if col in [2, 10, 12, 14, 15, 16]:
                    cell.number_format = '#,##0.00'

                # Aplicando cores conforme as lógicas
                if col == 10 and liquido_cell_fill:  # Valor Líquido
                    cell.fill = liquido_cell_fill
                elif col == 14 and pis_fill:  # PIS
                    cell.fill = pis_fill
                elif col == 15 and cofins_fill:  # COFINS
                    cell.fill = cofins_fill
                elif col == 17:  # Emissão de NFe
                    cell.fill = emissao_fill

        # -----------------------------
        # 5) Ajuste automático das colunas
        # -----------------------------
        for column_cells in ws.columns:
            length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if cell.value:
                        # Considera o tamanho do texto
                        cell_length = len(str(cell.value))
                        if cell_length > length:
                            length = cell_length
                except:
                    pass

            # Define largura mínima e máxima
            adjusted_width = min(max(length + 2, 10), 50)
            ws.column_dimensions[column].width = adjusted_width

        # Ajustes específicos para colunas importantes
        ws.column_dimensions['A'].width = 25  # Descrição
        ws.column_dimensions['B'].width = 12  # Valor
        ws.column_dimensions['C'].width = 12  # NF Filha
        ws.column_dimensions['J'].width = 15  # Valor Líquido
        ws.column_dimensions['M'].width = 18  # Classificação
        ws.column_dimensions['Q'].width = 18  # Emissão NFe

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="relatorio.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
